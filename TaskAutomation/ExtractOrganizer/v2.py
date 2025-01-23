import os
from PyPDF2 import PdfReader
from openpyxl import load_workbook

def get_values_from_xlsx(file_path: str, sheet_name: str = "Dados") -> list:
    contas, renomear_extrato = list(), list()

    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=3, values_only=True):
        conta = row[3]
        renomear_extrato_val = row[4]
        
        if conta is not None:
            contas.append(str(conta).strip())
        if renomear_extrato_val is not None:
            renomear_extrato.append(str(renomear_extrato_val).strip())

    return contas, renomear_extrato

def pdf_to_txt(path: str) -> str:
    text = ""
    pdf = PdfReader(path)
    for page in pdf.pages:
        text += page.extract_text() or ""
    return text

def listar_arquivos(pasta) -> list:
    if os.path.isdir(pasta):
        arquivos = os.listdir(pasta)
        arquivos = [arq for arq in arquivos if os.path.isfile(os.path.join(pasta, arq))]
        return arquivos
    else:
        print("O caminho fornecido não é uma pasta.")
        return []

excel_path = "/home/notebook-0359/Downloads/renato01.xlsx"
contas, nomes = get_values_from_xlsx(excel_path)

final_folder = "/home/notebook-0359/Downloads/ExtratoRenomear/"
arquivos = listar_arquivos(final_folder)

for arquivo in arquivos:
    path = final_folder + arquivo
    
    for (conta, nome) in zip(contas, nomes):
        if conta in pdf_to_txt(path):
            os.rename(path, f"{final_folder}{nome}.pdf")
            break